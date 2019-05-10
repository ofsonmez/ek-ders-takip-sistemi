from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from .models import Dersler, BolumBaskani
from .forms import *
from openpyxl import load_workbook
from datetime import datetime
import xlrd

now = timezone.now()

#Djanngonun local format la sıkıntıları olduğu için teşekkür ediyoruz :))
GG_DJANGO = {
    '01': 'Ocak', '02': 'Şubat', '03': 'Mart', '04': 'Nisan', '05': 'Mayıs', '06':'Haziran',
    '07': 'Temmuz', '08': 'Ağustos', '09': 'Eylül', '10':'Ekim', '11': 'Kasım',
    '12': 'Aralık'
}


#----------Ortak Fonksiyonlar----------#
#Ders kodundan bosluk silmek için 
def koddanBoslukSil(kod):
	if kod[3] == ' ':
		kod = kod.replace(' ', '')
	return kod

#----------Admin Donemlik Ders Programını Yüklediginde Calisicak Fonksiyon----------#   
#Databaseye'e eklemek için, calistiginda database'i sifirlar
def donemFileHandle(bas_satir, son_satir, file):
	Dersler.objects.all().delete()
	sayfa = file.sheet_by_index(0)
	for x in range(bas_satir, son_satir):
		if sayfa.cell_value(x, 1) != '':
			anlikdersiVeren = kelime_manipule(kelime_manipule(sayfa.cell_value(x, 1), "I", "ı"), "İ", "i").lower()
		if sayfa.cell_value(x, 2) == '':
			continue
		anlikDersKodu = koddanBoslukSil(sayfa.cell_value(x, 2))
		anlikDersAdı = sayfa.cell_value(x, 3)
		anlikKredi = sayfa.cell_value(x, 4)
		anlikDers = Dersler(dersiVeren=anlikdersiVeren, dersAdı=anlikDersAdı, dersKodu=anlikDersKodu, dersKredisi=anlikKredi)
		anlikDers.save()


#----------Kullanıcı Excel'i Yüklediginde Calisicek Fonksiyonlar----------#
#Excel dosyasında ilk boş satırı bulmak için gerekli fonk

def kelime_manipule(kelime, char1, char2):
	new = list(kelime)
	for x in range(len(new)):
		if new[x] == char1:
			new[x] = char2
	return "".join(new)

def bos_row(row):
	for x in range(len(row)):
		if row[x] != '':
			return False
	return True


#Dersler dizisinde excelden okunan ders varmı diye bakan fonk
#Ders varsa index numarasını, yoksa -1 dondurur
def search(cikti, anlik):
	for x in range(len(cikti)):
		if cikti[x]["kod"] == anlik["kod"] and cikti[x]["gun"] == anlik["gun"]:
			return x
	return -1

def search_kredi(cikti, anlik):
	for x in range(len(cikti)):
		if cikti[x]["kod"] == anlik["kod"]:
			return False
	return True        

def kredi_cozumle(cikti):
	carpim = 0
	for x in range(len(cikti)):
		flag = 0
		for y in range(len(cikti[x]["kredi"])):
			if cikti[x]["kredi"][y] == "*":
				carpim = int(cikti[x]["kredi"][y-1]) 
				flag = 1
			if flag == 1 and cikti[x]["kredi"][y].isnumeric():
				cikti[x]["kredi"][y] = str(int(cikti[x]["kredi"][y]) * carpim)


def kredi_hazirla(cikti):
	for x in range(len(cikti)):
		for y in range(len(cikti[x]["kredi"])):
			if cikti[x]["kredi"][y] == "*":
				del cikti[x]["kredi"][y-1]
				break

def kredi_toparla(kredi):
	for x in kredi:
		for y in x["kredi"][:]:
			if not y.isnumeric():
				x["kredi"].remove(y)

def kredi_kontrol(kredi,cikti):
	toplam_saat = 0
	toplam_kredi = 0
	for i in kredi:
		for j in cikti:
			if i['kod'] == j["kod"]:
				toplam_saat+=j["saat"]
		for k in i['kredi']:
			toplam_kredi+=int(k)
		if toplam_kredi != toplam_saat:
			return i["kod"]
		toplam_saat = 0
		toplam_kredi = 0
	return 'No Problem'

def bilgilendir(ders, kredi, cikti):
	for x in kredi:
		if ders["kod"] == x["kod"]:
			for y in range(len(x["kredi"])):
				if str(ders["saat"]) == x["kredi"][y]:
					x["kredi"][y] = "0"
					if y == 0:
						return "T"
					else:
						return "U"
			for y in range(len(x["kredi"])):
				if ders["saat"] < int(x["kredi"][y]):
					x["kredi"][y] = str(int(x["kredi"][y]) - ders["saat"])
					if y == 0:
						return "T"
					else:
						return "U"
			for y in range(len(x["kredi"])):
				if ders["saat"] > int(x["kredi"][y]):
					new_ders = ders.copy()
					new_ders["saat"] = ders["saat"] - int(x["kredi"][y])
					cikti.append(new_ders)
					ders["saat"] = int(x["kredi"][y])
					x["kredi"][y] = "0"
					if y == 0:
						return "T"
					else:
						return "U"


#Report excelden dersleri çekip diziye atan fonk
def dersleri_cek(sayfa, donguBas, donguSon, Gun):
	cikti = []
	anlik = {
		"kod" : "",
		"dersAdı": "",
		"gun": "",
		"saat": 1,
		"ikinciOgretim": False,
	}    
	for x in range(donguBas, donguSon):
		for y in range(1, sayfa.ncols, 7):
			if not sayfa.cell_value(x, y-1).isnumeric():
			#if sayfa.cell_value(x, y-1) != "1" and sayfa.cell_value(x, y-1) != "5":
				continue
			anlik["kod"] = sayfa.cell_value(x, y)
			anlik["gun"] = sayfa.cell_value(Gun, y-1)
			index = search(cikti, anlik)
			if index != -1:
				cikti[index]["saat"] += 1
			else:
				cikti.append(anlik.copy())    
	return cikti


#Excel dosyası üzerinde hangi fonksiyonların calisması gerektigini belirleyen ana fonksiyon
def excelFileHandle(bas_tarih, son_tarih, file, dersler,request):
	user = request.user
	cikti = []
	kredi = []
	anlik_kredi = {
		"kod" : "",
		"kredi": [],
	}
	sayfa = file.sheet_by_index(0)
	ilk_son = 0
	
	for x in range(3, sayfa.nrows):
		row = sayfa.row_values(x)
		if bos_row(row):
			ilk_son = x+4
			break
	
	if ilk_son == 0:
		cikti = dersleri_cek(sayfa=sayfa, donguBas=3, donguSon=sayfa.nrows, Gun=1)
	else:
		cikti = dersleri_cek(sayfa=sayfa, donguBas=3, donguSon=ilk_son, Gun=1)
		cikti += dersleri_cek(sayfa=sayfa, donguBas=ilk_son, donguSon=sayfa.nrows, Gun=ilk_son)

	for ders in cikti:
		ders["kod"] = koddanBoslukSil(ders["kod"])
		anlik_kredi["kod"] = ders["kod"]	
		try:
			gelen_ders = dersler.get(dersKodu=ders["kod"])
		except ObjectDoesNotExist:
			return render(request,'error.html',{'message': 'Yüklediginiz Formla Sistemdeki Dersleriniz Uyuşmuyor, Derslerinizi Duzenleyin.'})
		ders["dersAdı"] = gelen_ders.dersAdı
		if gelen_ders.ikinciOgretim:
			ders["ikinciOgretim"] = True
		if search_kredi(kredi, anlik_kredi):
			for y in range(len(gelen_ders.dersKredisi)):
				anlik_kredi["kredi"].append(gelen_ders.dersKredisi[y])
			kredi.append(anlik_kredi.copy())
			anlik_kredi["kredi"] = []

	kredi_cozumle(kredi)
	kredi_hazirla(kredi)
	kredi_toparla(kredi)
	kontrol = kredi_kontrol(kredi,cikti)
	
	if  kontrol != 'No Problem':
		return render(request, 'error.html', {'message': f'{kontrol} Kodlu Dersinizin Kredisi Yanlış Girilmiş Lütfen Düzeltin.'})
	
	#for x in kredi:
	#	print(x)
	#for ders in cikti:
	#	print(ders)
	#print(kredi_kontrol(kredi,cikti))

	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename=Ek-Ders-Ücret-Formu.xlsx'
	wb = load_workbook(filename='bos-ek-ders-ücret-formu.xlsx')
	ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	
	#  İsim-Soyisim
	ws.cell(row = 3, column = 12).value = "Dr.Öğr.Üyesi" + " " + user.first_name + " " + user.last_name

	#  Zorunlu ders miktarı
	ws.cell(row = 5, column = 12).value = 10

	#  Tc kimlik nosu
	ws.cell(row = 7, column = 12).value = user.username

	#  Yıl
	ws.cell(row = 4, column = 52).value = now.year

	#  Ay
	ws.cell(row = 5, column = 52).value = GG_DJANGO[str(now.strftime('%m'))]
	#  Hafta sayısı
	delta = son_tarih - bas_tarih
	ws.cell(row = 6, column = 52).value = int((delta.days) / 7)

	#  Normal Öğretim
	ws.cell(row = 6, column = 36).value = 8

	#  Başlangıç Tarihi
	ws.cell(row = 7, column = 46).value = bas_tarih

	#  Bitiş Tarihi
	ws.cell(row = 8, column = 46).value = son_tarih

	#Toplam DersYuku
	ws.cell(row = 69, column= 3).value = user.profile.dersYuku
	
	try:
		bolum_baskani = BolumBaskani.objects.get(bolum=user.profile.bolum).ad
	except BolumBaskani.DoesNotExist:
		bolum_baskani = None
	
	ws.cell(row = 80, column = 33).value = bolum_baskani

	sayac = 25
	ikinci_sayac = 48
	for ders in cikti:
		tur = bilgilendir(ders=ders,kredi=kredi,cikti=cikti)
		if ders["ikinciOgretim"]:
			ucuncu_sayac = ikinci_sayac
			ikinci_sayac+=1
		else:
			ucuncu_sayac = sayac
			sayac+=1 
		ws.cell(row=ucuncu_sayac, column = 16).value = ders["kod"]
		ws.cell(row=ucuncu_sayac, column = 50).value = ders["gun"]
		ws.cell(row=ucuncu_sayac, column = 20).value = ders["dersAdı"]
		ws.cell(row=ucuncu_sayac, column = 12).value = tur
		ws.cell(row=ucuncu_sayac, column = 2).value = kelime_manipule(user.profile.fakulte, "i", "İ").upper()
		if tur == "T":
			ws.cell(row=ucuncu_sayac, column = 38).value = ders["saat"]
		if tur == "U":
			ws.cell(row=ucuncu_sayac, column = 41).value = ders["saat"]
	#for x in kredi:
	#	print(x)
	wb.save(response)
	return response   